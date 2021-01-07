import requests
import re
from openpyxl import load_workbook, Workbook
import threading
from requests.packages.urllib3.exceptions import InsecureRequestWarning

requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

wb = Workbook()
wb.save('url.xlsx')
wb = load_workbook(r'url.xlsx')
sheet = wb.active
f = open('url.txt', 'r')
url = f.readlines()
thread_list = []


def scan_url(i):
    res = re.match('http://|https://', i)
    if res == None:
        url = 'http://' + i
    else:
        url = i
    try:
        if 'api' not in url:
            r = requests.get(url=url, timeout=3, allow_redirects=True, verify=False)
            r.encoding = r.apparent_encoding
            title = re.findall(r'<title>(.*)</title>', r.text)
            if (r.status_code != 403) and (r.status_code != 404) and (r.status_code != 406) and (
                    r.status_code != 502) and (title != '') and (title[0] != 'Welcome to OpenResty!') and (
                    title[0] != 'Welcome to nginx!'):
                data = url, title[0]
                res_tup = (data)
                print('\033[0;31m[+]\033[0m' + '\033[0;31m' + url + '\033[0m' + '   ' + '\033[0;31m' + title[
                    0] + '\033[0m')
                sheet.append(data)
    except:
        print('[+]URL ERROR：' + url)


print('''
\033[0;31m

   __  ______  __       _____ _________    _   __
  / / / / __ \/ /      / ___// ____/   |  / | / /
 / / / / /_/ / /       \__ \/ /   / /| | /  |/ / 
/ /_/ / _, _/ /___    ___/ / /___/ ___ |/ /|  /  
\____/_/ |_/_____/   /____/\____/_/  |_/_/ |_/   

            Author By Sam


\033[0m  
''')

for i in url:
    i = i.strip()
    t = threading.Thread(target=scan_url(i), args=i)
    thread_list.append(t)
    t.start()
for t in thread_list:
    t.join()
wb.save('url.xlsx')
